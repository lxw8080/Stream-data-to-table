#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
客户类型标记工具
用于自动标记2025年的新增客户订单和存量客户订单
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import argparse
import sys


class CustomerTypeMarker:
    """客户类型标记器"""
    
    # 颜色定义
    COLOR_NEW_CUSTOMER = "C6EFCE"  # 浅绿色 - 新增客户
    COLOR_EXISTING_CUSTOMER = "FFEB9C"  # 浅黄色 - 存量客户
    
    def __init__(self, file_path, sheet_name='Sheet1'):
        """
        初始化标记器
        
        Args:
            file_path: Excel文件路径
            sheet_name: 工作表名称，默认为'Sheet1'
        """
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.df = None
        self.wb = None
        self.ws = None
        
    def load_data(self):
        """加载Excel数据"""
        try:
            # 使用pandas读取数据，跳过第一行标题行
            self.df = pd.read_excel(
                self.file_path, 
                sheet_name=self.sheet_name,
                header=1  # 第2行作为表头
            )
            
            # 使用openpyxl加载工作簿用于格式化
            self.wb = load_workbook(self.file_path)
            self.ws = self.wb[self.sheet_name]
            
            print(f"✓ 成功加载文件: {self.file_path}")
            print(f"  工作表: {self.sheet_name}")
            print(f"  数据行数: {len(self.df)}")
            
            return True
        except Exception as e:
            print(f"✗ 加载文件失败: {e}")
            return False
    
    def validate_columns(self):
        """验证必要的列是否存在"""
        # 获取实际的列名
        columns = self.df.columns.tolist()
        print(f"\n当前列名: {columns}")
        
        # 检查列索引（C列=2, G列=6, T列=19，从0开始）
        if len(columns) < 20:
            print(f"✗ 列数不足，当前只有 {len(columns)} 列，需要至少20列")
            return False
        
        print(f"✓ 列验证通过")
        return True
    
    def mark_customer_types(self):
        """标记客户类型"""
        try:
            # 获取列（pandas使用0-based索引）
            # C列(索引2): 订单创建日期
            # G列(索引6): 客户身份证号
            # T列(索引19): 客户类型标记
            
            date_col_idx = 2  # C列
            id_col_idx = 6    # G列
            mark_col_idx = 19 # T列
            
            # 获取列名
            date_col = self.df.columns[date_col_idx]
            id_col = self.df.columns[id_col_idx]
            mark_col = self.df.columns[mark_col_idx]
            
            print(f"\n使用的列:")
            print(f"  订单创建日期列 (C列): {date_col}")
            print(f"  客户身份证号列 (G列): {id_col}")
            print(f"  客户类型标记列 (T列): {mark_col}")
            
            # 转换日期列为datetime类型
            self.df[date_col] = pd.to_datetime(self.df[date_col], errors='coerce')
            
            # 定义2025年的起始日期
            year_2025_start = pd.Timestamp('2025-01-01')
            
            # 获取2025年之前出现过的所有身份证号（去重）
            pre_2025_customers = set(
                self.df[self.df[date_col] < year_2025_start][id_col].dropna().unique()
            )
            
            print(f"\n2025年之前的客户数量: {len(pre_2025_customers)}")
            
            # 初始化标记列
            self.df[mark_col] = ''
            
            new_count = 0
            existing_count = 0
            
            # 遍历每一行进行标记
            for idx, row in self.df.iterrows():
                order_date = row[date_col]
                customer_id = row[id_col]
                
                # 跳过日期或身份证号为空的行
                if pd.isna(order_date) or pd.isna(customer_id):
                    continue
                
                # 2025年之前的订单，标记为"存量"
                if order_date < year_2025_start:
                    self.df.at[idx, mark_col] = '存量'
                    existing_count += 1
                else:
                    # 2025年及以后的订单
                    if customer_id in pre_2025_customers:
                        # 该客户在2025年之前出现过，标记为"存量"
                        self.df.at[idx, mark_col] = '存量'
                        existing_count += 1
                    else:
                        # 该客户在2025年之前未出现过，标记为"新增"
                        self.df.at[idx, mark_col] = '新增'
                        new_count += 1
            
            print(f"\n标记完成:")
            print(f"  新增客户订单: {new_count} 条")
            print(f"  存量客户订单: {existing_count} 条")
            
            return True
        except Exception as e:
            print(f"✗ 标记客户类型失败: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def apply_formatting(self):
        """应用单元格格式（颜色）"""
        try:
            # T列在Excel中是第20列
            mark_col_letter = 'T'
            
            # 创建填充样式
            fill_new = PatternFill(start_color=self.COLOR_NEW_CUSTOMER, 
                                   end_color=self.COLOR_NEW_CUSTOMER, 
                                   fill_type='solid')
            fill_existing = PatternFill(start_color=self.COLOR_EXISTING_CUSTOMER, 
                                       end_color=self.COLOR_EXISTING_CUSTOMER, 
                                       fill_type='solid')
            
            # 从第3行开始（跳过标题行和表头行）
            start_row = 3
            
            for idx, row in self.df.iterrows():
                excel_row = start_row + idx
                mark_col_idx = 19  # T列索引
                mark_value = row[self.df.columns[mark_col_idx]]
                
                cell = self.ws[f'{mark_col_letter}{excel_row}']
                cell.value = mark_value
                
                # 根据标记值应用颜色
                if mark_value == '新增':
                    cell.fill = fill_new
                elif mark_value == '存量':
                    cell.fill = fill_existing
            
            print(f"✓ 格式应用完成")
            return True
        except Exception as e:
            print(f"✗ 应用格式失败: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def save_file(self, output_path=None):
        """保存文件"""
        try:
            if output_path is None:
                output_path = self.file_path
            
            self.wb.save(output_path)
            print(f"✓ 文件已保存: {output_path}")
            return True
        except Exception as e:
            print(f"✗ 保存文件失败: {e}")
            return False
    
    def process(self, output_path=None):
        """执行完整的处理流程"""
        print("=" * 60)
        print("客户类型标记工具")
        print("=" * 60)
        
        # 1. 加载数据
        if not self.load_data():
            return False
        
        # 2. 验证列
        if not self.validate_columns():
            return False
        
        # 3. 标记客户类型
        if not self.mark_customer_types():
            return False
        
        # 4. 应用格式
        if not self.apply_formatting():
            return False
        
        # 5. 保存文件
        if not self.save_file(output_path):
            return False
        
        print("\n" + "=" * 60)
        print("处理完成！")
        print("=" * 60)
        print(f"\n颜色说明:")
        print(f"  🟢 浅绿色 - 新增客户")
        print(f"  🟡 浅黄色 - 存量客户")
        
        return True


def find_excel_file():
    """在当前目录查找Excel文件"""
    import glob
    excel_files = glob.glob('*.xlsx')
    # 排除临时文件和测试文件
    excel_files = [
        f for f in excel_files
        if not f.startswith('~') and not f.startswith('测试')
    ]
    return excel_files


def main():
    """命令行入口"""
    parser = argparse.ArgumentParser(
        description='自动标记2025年的新增客户订单和存量客户订单',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例用法:
  python mark_customer_type.py 租机登记表.xlsx
  python mark_customer_type.py 租机登记表.xlsx -o 标记结果.xlsx
  python mark_customer_type.py 租机登记表.xlsx -s Sheet1

如果不指定文件，将自动查找当前目录的Excel文件
        """
    )

    parser.add_argument(
        'file', nargs='?',
        help='Excel文件路径（可选，不指定则自动查找）'
    )
    parser.add_argument(
        '-s', '--sheet', default='Sheet1',
        help='工作表名称（默认: Sheet1）'
    )
    parser.add_argument(
        '-o', '--output',
        help='输出文件路径（默认: 覆盖原文件）'
    )

    args = parser.parse_args()

    # 如果没有指定文件，尝试自动查找
    file_path = args.file
    if not file_path:
        excel_files = find_excel_file()
        if not excel_files:
            print("✗ 错误: 当前目录未找到Excel文件")
            print("请指定文件路径，例如:")
            print("  python mark_customer_type.py 租机登记表.xlsx")
            sys.exit(1)
        elif len(excel_files) == 1:
            file_path = excel_files[0]
            print(f"自动选择文件: {file_path}\n")
        else:
            print("当前目录有多个Excel文件，请选择:")
            for i, f in enumerate(excel_files, 1):
                print(f"  {i}. {f}")
            try:
                prompt = "\n请输入序号 (1-{}): ".format(len(excel_files))
                choice = int(input(prompt))
                if 1 <= choice <= len(excel_files):
                    file_path = excel_files[choice - 1]
                    print(f"已选择: {file_path}\n")
                else:
                    print("✗ 无效的选择")
                    sys.exit(1)
            except (ValueError, KeyboardInterrupt):
                print("\n✗ 已取消")
                sys.exit(1)

    # 创建标记器并执行
    marker = CustomerTypeMarker(file_path, args.sheet)
    success = marker.process(args.output)

    sys.exit(0 if success else 1)


if __name__ == '__main__':
    main()
