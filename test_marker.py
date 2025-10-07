#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
客户类型标记工具测试脚本
用于验证工具功能是否正常
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime, timedelta
import random
import os


def generate_test_data():
    """生成测试数据"""
    print("生成测试数据...")
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    
    # 第1行：标题行
    ws['A1'] = '租机登记表'
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:T1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 第2行：表头行
    headers = [
        '序号', '订单编号', '订单创建日期', '客户姓名', '联系电话',
        '客户地址', '客户身份证号', '设备型号', '设备编号', '租赁开始日期',
        '租赁结束日期', '月租金', '押金', '支付方式', '订单状态',
        '业务员', '备注', '创建人', '更新时间', '客户类型'
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    
    # 生成测试数据
    # 模拟100个客户，其中70个是2025年之前的老客户
    old_customers = []
    for i in range(70):
        old_customers.append(f'3301{1980+i%40:04d}{random.randint(10,12):02d}{random.randint(10,28):02d}{random.randint(1000,9999):04d}')
    
    # 30个新客户（仅在2025年出现）
    new_customers = []
    for i in range(30):
        new_customers.append(f'3302{1985+i%35:04d}{random.randint(1,12):02d}{random.randint(1,28):02d}{random.randint(1000,9999):04d}')
    
    # 生成订单数据
    row_num = 3
    order_id = 1000
    
    # 2024年的订单（使用老客户）
    print("  生成2024年订单...")
    base_date_2024 = datetime(2024, 1, 1)
    for i in range(50):
        customer_id = random.choice(old_customers)
        order_date = base_date_2024 + timedelta(days=random.randint(0, 364))
        
        ws.cell(row=row_num, column=1, value=row_num-2)  # 序号
        ws.cell(row=row_num, column=2, value=f'ORD{order_id}')  # 订单编号
        ws.cell(row=row_num, column=3, value=order_date)  # 订单创建日期
        ws.cell(row=row_num, column=4, value=f'客户{i+1}')  # 客户姓名
        ws.cell(row=row_num, column=5, value=f'138{random.randint(10000000,99999999)}')  # 联系电话
        ws.cell(row=row_num, column=6, value=f'杭州市西湖区XX路{i+1}号')  # 客户地址
        ws.cell(row=row_num, column=7, value=customer_id)  # 客户身份证号
        ws.cell(row=row_num, column=8, value=f'型号-{random.choice(["A", "B", "C"])}')  # 设备型号
        ws.cell(row=row_num, column=9, value=f'DEV{random.randint(1000,9999)}')  # 设备编号
        ws.cell(row=row_num, column=10, value=order_date)  # 租赁开始日期
        ws.cell(row=row_num, column=11, value=order_date + timedelta(days=365))  # 租赁结束日期
        ws.cell(row=row_num, column=12, value=random.randint(200, 500))  # 月租金
        ws.cell(row=row_num, column=13, value=random.randint(1000, 3000))  # 押金
        ws.cell(row=row_num, column=14, value=random.choice(['支付宝', '微信', '银行转账']))  # 支付方式
        ws.cell(row=row_num, column=15, value='已完成')  # 订单状态
        ws.cell(row=row_num, column=16, value=f'业务员{random.randint(1,5)}')  # 业务员
        ws.cell(row=row_num, column=17, value='')  # 备注
        ws.cell(row=row_num, column=18, value='系统')  # 创建人
        ws.cell(row=row_num, column=19, value=datetime.now())  # 更新时间
        # T列（第20列）留空，等待工具填充
        
        row_num += 1
        order_id += 1
    
    # 2025年的订单（混合老客户和新客户）
    print("  生成2025年订单（老客户）...")
    base_date_2025 = datetime(2025, 1, 1)
    for i in range(30):
        customer_id = random.choice(old_customers)  # 老客户
        order_date = base_date_2025 + timedelta(days=random.randint(0, 90))
        
        ws.cell(row=row_num, column=1, value=row_num-2)
        ws.cell(row=row_num, column=2, value=f'ORD{order_id}')
        ws.cell(row=row_num, column=3, value=order_date)
        ws.cell(row=row_num, column=4, value=f'客户{i+51}')
        ws.cell(row=row_num, column=5, value=f'138{random.randint(10000000,99999999)}')
        ws.cell(row=row_num, column=6, value=f'杭州市西湖区XX路{i+51}号')
        ws.cell(row=row_num, column=7, value=customer_id)
        ws.cell(row=row_num, column=8, value=f'型号-{random.choice(["A", "B", "C"])}')
        ws.cell(row=row_num, column=9, value=f'DEV{random.randint(1000,9999)}')
        ws.cell(row=row_num, column=10, value=order_date)
        ws.cell(row=row_num, column=11, value=order_date + timedelta(days=365))
        ws.cell(row=row_num, column=12, value=random.randint(200, 500))
        ws.cell(row=row_num, column=13, value=random.randint(1000, 3000))
        ws.cell(row=row_num, column=14, value=random.choice(['支付宝', '微信', '银行转账']))
        ws.cell(row=row_num, column=15, value='进行中')
        ws.cell(row=row_num, column=16, value=f'业务员{random.randint(1,5)}')
        ws.cell(row=row_num, column=17, value='')
        ws.cell(row=row_num, column=18, value='系统')
        ws.cell(row=row_num, column=19, value=datetime.now())
        
        row_num += 1
        order_id += 1
    
    print("  生成2025年订单（新客户）...")
    for i in range(20):
        customer_id = random.choice(new_customers)  # 新客户
        order_date = base_date_2025 + timedelta(days=random.randint(0, 90))
        
        ws.cell(row=row_num, column=1, value=row_num-2)
        ws.cell(row=row_num, column=2, value=f'ORD{order_id}')
        ws.cell(row=row_num, column=3, value=order_date)
        ws.cell(row=row_num, column=4, value=f'客户{i+81}')
        ws.cell(row=row_num, column=5, value=f'138{random.randint(10000000,99999999)}')
        ws.cell(row=row_num, column=6, value=f'杭州市西湖区XX路{i+81}号')
        ws.cell(row=row_num, column=7, value=customer_id)
        ws.cell(row=row_num, column=8, value=f'型号-{random.choice(["A", "B", "C"])}')
        ws.cell(row=row_num, column=9, value=f'DEV{random.randint(1000,9999)}')
        ws.cell(row=row_num, column=10, value=order_date)
        ws.cell(row=row_num, column=11, value=order_date + timedelta(days=365))
        ws.cell(row=row_num, column=12, value=random.randint(200, 500))
        ws.cell(row=row_num, column=13, value=random.randint(1000, 3000))
        ws.cell(row=row_num, column=14, value=random.choice(['支付宝', '微信', '银行转账']))
        ws.cell(row=row_num, column=15, value='进行中')
        ws.cell(row=row_num, column=16, value=f'业务员{random.randint(1,5)}')
        ws.cell(row=row_num, column=17, value='')
        ws.cell(row=row_num, column=18, value='系统')
        ws.cell(row=row_num, column=19, value=datetime.now())
        
        row_num += 1
        order_id += 1
    
    # 调整列宽
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['T'].width = 12
    
    # 保存文件
    test_file = '测试_租机登记表.xlsx'
    wb.save(test_file)
    print(f"✓ 测试数据已生成: {test_file}")
    print(f"  总订单数: {row_num - 3}")
    print(f"  2024年订单: 50条（应全部标记为'存量'）")
    print(f"  2025年老客户订单: 30条（应标记为'存量'）")
    print(f"  2025年新客户订单: 20条（应标记为'新增'）")
    
    return test_file


def run_test():
    """运行测试"""
    print("=" * 60)
    print("客户类型标记工具 - 测试脚本")
    print("=" * 60)
    print()
    
    # 生成测试数据
    test_file = generate_test_data()
    
    print("\n" + "=" * 60)
    print("测试数据生成完成！")
    print("=" * 60)
    print("\n请运行以下命令测试工具：")
    print(f"  python mark_customer_type.py \"{test_file}\" -o \"测试结果.xlsx\"")
    print("\n预期结果：")
    print("  - 2024年的50条订单应标记为'存量'（浅黄色）")
    print("  - 2025年老客户的30条订单应标记为'存量'（浅黄色）")
    print("  - 2025年新客户的20条订单应标记为'新增'（浅绿色）")
    print()


if __name__ == '__main__':
    run_test()

